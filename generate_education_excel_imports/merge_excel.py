import os
import openpyxl

def create_workbook_with_sheets(sheet_names, directory='sheets'):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet
    for name in sheet_names:
        wb.create_sheet(title=name[:-5])
    return wb

def copy_data_to_sheet(source_file, target_sheet):
    source_wb = openpyxl.load_workbook(source_file)
    source_sheet = source_wb.active
    for row in source_sheet.iter_rows(values_only=True):
        target_sheet.append(row)

def main(directory='sheets'):
    sheet_files = ['establishments.xlsx', 'classes.xlsx', 'teachers.xlsx', 'students.xlsx']
    wb = create_workbook_with_sheets(sheet_files, directory)
    for sheet_name, file_name in zip(wb.sheetnames, sheet_files):
        file_path = os.path.join(directory, file_name)
        copy_data_to_sheet(file_path, wb[sheet_name])
    wb.save(os.path.join(directory, 'education_import_generated.xlsx'))

if __name__ == "__main__":
    main()
